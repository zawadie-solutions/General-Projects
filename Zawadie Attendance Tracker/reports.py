import csv
from datetime import datetime
from pathlib import Path
from typing import Any

from attendance import calculate_hours
from settings import ATTENDANCE_HEADERS, EXPORTS_DIR, ensure_directories


def display_date(iso_date: str) -> str:
    try:
        return datetime.strptime(iso_date, "%Y-%m-%d").strftime("%d/%m/%Y")
    except ValueError:
        return iso_date


def report_rows(records: list[dict[str, Any]]) -> list[list[str]]:
    rows = []
    for record in records:
        rows.append(
            [
                display_date(record.get("attendance_date", "")),
                record.get("employee_id", ""),
                record.get("full_name", ""),
                record.get("gender", ""),
                record.get("department", ""),
                str(record.get("session_number", 1)),
                record.get("sign_in") or "",
                record.get("sign_out") or "",
                calculate_hours(record.get("sign_in"), record.get("sign_out")),
                record.get("reentry_reason") or "",
            ]
        )
    return rows


def export_csv(records: list[dict[str, Any]], path: Path | None = None) -> Path:
    ensure_directories()
    output_path = path or EXPORTS_DIR / _filename("attendance", "csv")
    with output_path.open("w", newline="", encoding="utf-8") as file:
        writer = csv.writer(file)
        writer.writerow(ATTENDANCE_HEADERS)
        writer.writerows(report_rows(records))
    return output_path


def export_excel(records: list[dict[str, Any]], path: Path | None = None) -> Path:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill
    from openpyxl.utils import get_column_letter

    ensure_directories()
    output_path = path or EXPORTS_DIR / _filename("attendance", "xlsx")
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Attendance"
    sheet.append(ATTENDANCE_HEADERS)
    for row in report_rows(records):
        sheet.append(row)
    for cell in sheet[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="1F2937")
    for column_cells in sheet.columns:
        max_length = max(len(str(cell.value or "")) for cell in column_cells)
        sheet.column_dimensions[get_column_letter(column_cells[0].column)].width = max(12, max_length + 2)
    workbook.save(output_path)
    return output_path


def export_pdf(records: list[dict[str, Any]], path: Path | None = None) -> Path:
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.lib.styles import getSampleStyleSheet
    from reportlab.platypus import Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle

    ensure_directories()
    output_path = path or EXPORTS_DIR / _filename("attendance", "pdf")
    doc = SimpleDocTemplate(str(output_path), pagesize=landscape(A4), rightMargin=24, leftMargin=24)
    styles = getSampleStyleSheet()
    story = [Paragraph("Attendance Report", styles["Title"]), Spacer(1, 12)]
    table = Table([ATTENDANCE_HEADERS] + report_rows(records), repeatRows=1)
    table.setStyle(
        TableStyle(
            [
                ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#1F2937")),
                ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
                ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                ("GRID", (0, 0), (-1, -1), 0.25, colors.HexColor("#CBD5E1")),
                ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#F8FAFC")]),
                ("FONTSIZE", (0, 0), (-1, -1), 8),
                ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
            ]
        )
    )
    story.append(table)
    doc.build(story)
    return output_path


def _filename(prefix: str, extension: str) -> str:
    stamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    return f"{prefix}_{stamp}.{extension}"
